import os
import json
import sqlite3
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

# Support deployment at subpath via environment variable
SUBPATH = os.getenv('APP_SUBPATH', '').strip('/')
if SUBPATH:
    SUBPATH = '/' + SUBPATH

# Path to the built presentation files
PRESENTATION_DIR = os.path.join(os.path.dirname(__file__), 'presentation')

app = Flask(__name__, static_folder='assets', static_url_path=f'{SUBPATH}/assets')
DATABASE = os.path.join(os.path.dirname(__file__), 'responses.db')


def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS responses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            delivery_failure TEXT,
            supermarket_time TEXT,
            cool_locker TEXT,
            failed_find TEXT,
            auto_availability TEXT,
            meal_prep TEXT,
            shopping_recs TEXT,
            hygiene_importance INTEGER,
            would_pay TEXT,
            gender TEXT,
            age_range TEXT,
            email TEXT,
            name TEXT,
            submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()


# Initialize database at startup (works with both local and Gunicorn)
try:
    init_db()
except Exception as e:
    print(f"Warning: Database initialization error: {e}", flush=True)


@app.route('/coolbox/assets/<path:filename>')
def presentation_assets(filename):
    return send_from_directory(os.path.join(PRESENTATION_DIR, 'assets'), filename)


@app.route('/coolbox/')
def presentation_index():
    return send_from_directory(PRESENTATION_DIR, 'index.html')


@app.route('/coolbox/landingpage')
def presentation_landingpage():
    return send_from_directory(PRESENTATION_DIR, 'index.html')


@app.route(f'{SUBPATH}/' if SUBPATH else '/')
def index():
    return render_template('form.html', subpath=SUBPATH)


@app.route(f'{SUBPATH}/dashboard' if SUBPATH else '/dashboard')
def dashboard():
    return render_template('dashboard.html', subpath=SUBPATH)


@app.route(f'{SUBPATH}/api/submit' if SUBPATH else '/api/submit', methods=['POST'])

def submit():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data provided'}), 400

    conn = get_db()
    conn.execute('''
        INSERT INTO responses (delivery_failure, supermarket_time, cool_locker,
            failed_find, auto_availability, meal_prep, shopping_recs,
            hygiene_importance, would_pay, gender, age_range, email, name)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data.get('delivery_failure', ''),
        data.get('supermarket_time', ''),
        data.get('cool_locker', ''),
        data.get('failed_find', ''),
        data.get('auto_availability', ''),
        data.get('meal_prep', ''),
        data.get('shopping_recs', ''),
        data.get('hygiene_importance', None),
        data.get('would_pay', ''),
        data.get('gender', ''),
        data.get('age_range', ''),
        data.get('email', '').strip(),
        data.get('name', '').strip()
    ))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': 'Response saved!'})


def yn_stats(conn, column):
    rows = conn.execute(f'''
        SELECT {column}, COUNT(*) as count
        FROM responses WHERE {column} != ''
        GROUP BY {column}
    ''').fetchall()
    return [{'label': r[column], 'count': r['count']} for r in rows]


# All known options per question (order matters for chart display)
ALL_OPTIONS = {
    'supermarket_time': [
        "That's a lot, I can't stand it anymore",
        "I accept it, but I'm open to becoming more efficient",
        "It is what it is...",
    ],
    'delivery_failure': ['Yes', 'No'],
    'cool_locker': ['Yes', 'No'],
    'meal_prep': ['Yes', 'Maybe', 'No'],
    'shopping_recs': ['Yes', 'No'],
    'would_pay': ['Yes', 'No'],
    'gender': ['Male', 'Female', 'Prefer not to say'],
    'age_range': ['18-25 years', '25-35 years', '36-50 years', '50+ years'],
}


def full_stats(conn, column):
    """Return stats for a column, ensuring all known options are present (even with count 0)."""
    rows = conn.execute(f'''
        SELECT {column}, COUNT(*) as count
        FROM responses WHERE {column} != ''
        GROUP BY {column}
    ''').fetchall()
    counts = {r[column]: r['count'] for r in rows}

    options = ALL_OPTIONS.get(column)
    if options:
        return [{'label': opt, 'count': counts.get(opt, 0)} for opt in options]
    # Fallback for questions without a known option list
    return [{'label': r[column], 'count': r['count']} for r in rows]


@app.route(f'{SUBPATH}/api/stats' if SUBPATH else '/api/stats')
def stats():
    conn = get_db()

    total = conn.execute('SELECT COUNT(*) as c FROM responses').fetchone()['c']

    delivery_failure = full_stats(conn, 'delivery_failure')

    supermarket_time = full_stats(conn, 'supermarket_time')

    cool_locker = full_stats(conn, 'cool_locker')
    failed_find = full_stats(conn, 'failed_find')
    auto_availability = full_stats(conn, 'auto_availability')

    meal_prep = full_stats(conn, 'meal_prep')

    shopping_recs = full_stats(conn, 'shopping_recs')

    hygiene = conn.execute('''
        SELECT AVG(hygiene_importance) as avg_val,
               MIN(hygiene_importance) as min_val,
               MAX(hygiene_importance) as max_val
        FROM responses WHERE hygiene_importance IS NOT NULL
    ''').fetchone()

    hygiene_dist = conn.execute('''
        SELECT hygiene_importance, COUNT(*) as count
        FROM responses WHERE hygiene_importance IS NOT NULL
        GROUP BY hygiene_importance ORDER BY hygiene_importance
    ''').fetchall()

    would_pay = full_stats(conn, 'would_pay')

    gender = full_stats(conn, 'gender')

    age_range = full_stats(conn, 'age_range')

    recent = conn.execute('''
        SELECT name, email, submitted_at FROM responses
        ORDER BY submitted_at DESC LIMIT 10
    ''').fetchall()

    # Count answered per question
    answered = {}
    for col in ['delivery_failure', 'supermarket_time', 'cool_locker', 'failed_find',
                'auto_availability', 'meal_prep', 'shopping_recs', 'would_pay',
                'gender', 'age_range', 'email', 'name']:
        r = conn.execute(f"SELECT COUNT(*) as c FROM responses WHERE {col} != ''").fetchone()
        answered[col] = r['c']
    r = conn.execute("SELECT COUNT(*) as c FROM responses WHERE hygiene_importance IS NOT NULL").fetchone()
    answered['hygiene_importance'] = r['c']

    conn.close()

    return jsonify({
        'total_responses': total,
        'answered': answered,
        'delivery_failure': delivery_failure,
        'supermarket_time': supermarket_time,
        'cool_locker': cool_locker,
        'failed_find': failed_find,
        'auto_availability': auto_availability,
        'meal_prep': meal_prep,
        'shopping_recs': shopping_recs,
        'hygiene_avg': round(hygiene['avg_val'], 1) if hygiene['avg_val'] else 0,
        'hygiene_dist': [{'label': str(r['hygiene_importance']), 'count': r['count']} for r in hygiene_dist],
        'would_pay': would_pay,
        'gender': gender,
        'age_range': age_range,
        'recent': [{'name': r['name'] or 'Anonymous', 'email': r['email'] or 'N/A', 'date': r['submitted_at']} for r in recent]
    })


@app.route(f'{SUBPATH}/api/export' if SUBPATH else '/api/export')
def export_excel():
    conn = get_db()
    rows = conn.execute('''
        SELECT delivery_failure, supermarket_time, cool_locker,
               meal_prep, shopping_recs, would_pay, gender, age_range, 
               email, name, submitted_at
        FROM responses ORDER BY submitted_at DESC
    ''').fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Form Responses"

    headers = [
        'Time in Supermarkets', 'Delivery Failure?', 'Cool Locker Pickup?',
        'Meal Prep Delivery?', 'Shopping Recommendations?', 'Would Pay?',
        'Gender', 'Age Range', 'Email', 'Name', 'Submitted At'
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="009000", end_color="009000", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row in enumerate(rows, 2):
        values = [
            row['supermarket_time'], row['delivery_failure'], row['cool_locker'],
            row['meal_prep'], row['shopping_recs'],
            row['would_pay'], row['gender'], row['age_range'], 
            row['email'], row['name'], row['submitted_at']
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    for col in range(1, len(headers) + 1):
        max_length = len(str(ws.cell(row=1, column=col).value))
        for row in range(2, len(rows) + 2):
            cell_val = ws.cell(row=row, column=col).value
            if cell_val:
                max_length = max(max_length, len(str(cell_val)))
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max_length + 4, 50)

    # Set auto filter to header row and all data
    if len(rows) > 0:
        last_col_letter = ws.cell(row=1, column=len(headers)).column_letter
        ws.auto_filter.ref = f"A1:{last_col_letter}{len(rows) + 1}"

    # Use BytesIO to avoid temp file issues
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f'fresh_groceries_responses_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=8080)
